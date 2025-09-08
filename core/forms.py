from django import forms
from django.utils.deconstruct import deconstructible
from django.core.exceptions import ValidationError
import pandas as pd
from openpyxl import load_workbook
from .models import Carteira
from datetime import datetime
import re
from .models import Acordo


@deconstructible
class ExcelFileValidator:
    def __call__(self, value):
        if not value:
            # Se o arquivo não for fornecido, nao tem nada a validar
            return

        if not value.name.endswith('.xlsx'):
            raise ValidationError('O arquivo deve ter a extensão .xlsx.')
        
        try:
            wb = load_workbook(value)
        except Exception as e:
            raise ValidationError(f'Erro ao carregar o arquivo Excel: {e}')

        value.seek(0)
        
        try:
            df = pd.read_excel(value)

            # Verificando colunas obrigatórias
            required_columns = ['Num Original do Contrato','Data Inicio Contrato', 'Data Final Contrato', 'CPF', 'Produto', 
                                'Data de Vencimento da Parcela', 'Num da Parcela', 'Valor da Parcela', 'Data Pagamento Parcela', 'Valor do Pagamento', 'Taxa Mensal Efetiva', 
                                'Taxa Mensal Nominal', 'Valor Mora', 'Valor Multa', 
                                'Valor Principal', 'Valor de Juros', 'Ultima Data Restricao', 'Inclusao Restricao', 'Orgao de Restricao', 
                                'Protesto', 'Data Protesto', 'Credito Financiado', 'Credito Liberado' ]
            missing_columns = [col for col in required_columns if col not in df.columns]

            if len(missing_columns) == 1 and 'Produto' in missing_columns:
                # A coluna 'Produto' é a única opcional, então se ela estiver ausente, não é um problema
                pass
            elif missing_columns:
                raise ValidationError(f'O arquivo não contém as colunas obrigatórias: {", ".join(missing_columns)}.')

            # Verificando linhas com valores não nulos nas colunas
            if df[required_columns].dropna().empty:
                raise ValidationError('O arquivo enviado não segue os padrões fornecidos pelo arquivo de exemplo')
            
            self.validate_contrato_column(df['Num Original do Contrato'])
            self.validate_data_inicio_contrato(df['Data Inicio Contrato'])
            self.validate_data_final_contrato(df['Data Final Contrato'])
            self.validate_cpf_format(df['CPF'])
            self.validate_data_vencimento_parcela(df['Data de Vencimento da Parcela'])
            self.validate_num_parcela(df['Num da Parcela'])
            self.validate_valor_parcela_format(df['Valor da Parcela'])
            self.validate_data_pagamento_parcela_column(df['Data Pagamento Parcela'])
            self.validate_valor_pagamento_format(df['Valor do Pagamento'])
            self.validate_taxa_mensal_efetiva(df['Taxa Mensal Efetiva'])
            self.validate_taxa_mensal_nominal(df['Taxa Mensal Nominal'])
            self.validate_valor_mora(df['Valor Mora'])
            self.validate_valor_multa(df['Valor Multa'])
            self.validate_valor_principal(df['Valor Principal'])
            self.validate_valor_juros(df['Valor de Juros'])
            self.validate_ultima_data_restricao_column(df['Ultima Data Restricao'])
            self.validate_inclusao_restricao_column(df['Inclusao Restricao'])
            self.validate_orgao_restricao_column(df['Orgao de Restricao'], df['Inclusao Restricao'])
            self.validate_protesto_column(df['Protesto'])
            self.validate_data_protesto_column(df['Data Protesto'], df['Protesto'])
            self.validate_credito_financiado_column(df['Credito Financiado'])
            self.validate_credito_liberado_column(df['Credito Liberado'])

        except pd.errors.ParserError as e:
            raise ValidationError(f'Erro ao analisar o arquivo Excel: {e}')

        value.seek(0)  # Volta ao início do arquivo para que possa ser salvo no modelo
        
    def validate_contrato_column(self, contrato_column):
        # Validar se a coluna 'Num Original do Contrato' contém apenas letras e/ou números
        for valor in contrato_column:
            if not any(char.isalnum() for char in str(valor)):
                raise ValidationError('A coluna Num Original do Contrato deve conter apenas letras e/ou números.')
            
    def validate_data_inicio_contrato(self, data_inicio_contrato_column):
    # Validar se a coluna 'Data Inicio Contrato' contém datas no formato 'DD/MM/AAAA' ou 'YYYY-MM-DD HH:MM:SS' ou 'DD/MM/YY'
        for valor in data_inicio_contrato_column:
            try:
            # Tenta primeiro o formato 'DD/MM/AAAA'
                datetime.strptime(str(valor), '%d/%m/%Y')
            except ValueError:
                try:
                # Se falhar, tenta o formato padrão do Pandas 'YYYY-MM-DD HH:MM:SS'
                    datetime.strptime(str(valor), '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                    # Se falhar novamente, tenta o formato com ano de dois dígitos 'DD/MM/YY'
                        datetime.strptime(str(valor), '%d/%m/%y')
                    except ValueError:
                    # Se ainda falhar, levanta a exceção
                        raise ValidationError('A coluna Data Inicio Contrato deve conter datas no formato DD/MM/AAAA ou YYYY-MM-DD HH:MM:SS ou DD/MM/YY.')
            
    def validate_data_final_contrato(self, data_final_contrato_column):
        # Validar se a coluna 'Data Final Contrato' contém datas no formato 'DD/MM/AA'
        for valor in data_final_contrato_column:
            try:
            # Tenta primeiro o formato 'DD/MM/AAAA'
                datetime.strptime(str(valor), '%d/%m/%Y')
            except ValueError:
                try:
                # Se falhar, tenta o formato padrão do Pandas 'YYYY-MM-DD HH:MM:SS'
                    datetime.strptime(str(valor), '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                    # Se falhar novamente, tenta o formato com ano de dois dígitos 'DD/MM/YY'
                        datetime.strptime(str(valor), '%d/%m/%y')
                    except ValueError:
                    # Se ainda falhar, levanta a exceção
                        raise ValidationError('A coluna Data Final Contrato deve conter datas no formato DD/MM/AAAA ou YYYY-MM-DD HH:MM:SS ou DD/MM/YY.')
    
    def validar_cpf(self,cpf):
    # Remover caracteres não numéricos
        cpf_numerico = ''.join(char for char in str(cpf) if char.isdigit())

    # Verificar se o CPF tem 11 dígitos
        if not cpf_numerico.isdigit() or len(cpf_numerico) != 11:
            return False

    # Calcular os dígitos verificadores
        cpf_digits = [int(digit) for digit in cpf_numerico]

        soma1 = sum(digit * weight for digit, weight in zip(cpf_digits[:9], range(10, 1, -1)))
        resto1 = soma1 % 11
        digito_verificador1 = 0 if resto1 < 2 else 11 - resto1

        soma2 = sum(digit * weight for digit, weight in zip(cpf_digits[:10], range(11, 1, -1)))
        resto2 = soma2 % 11
        digito_verificador2 = 0 if resto2 < 2 else 11 - resto2

    # Verificar se os dígitos verificadores estão corretos
        if cpf_digits[9] != digito_verificador1 or cpf_digits[10] != digito_verificador2:
            return False

        return True

            
    def validate_cpf_format(self, cpf_column):
    # Validar se a coluna 'CPF' contém CPFs no formato xxx.xxx.xxx-xx
        for cpf_str in cpf_column:
            if not self.validar_cpf(cpf_str):
                raise ValidationError('A coluna CPF deve conter um CPF válido e no formato XXX.XXX.XXX-XX.')
            
            
    def validate_data_vencimento_parcela(self, data_vencimento_parcela_column):
        # Validar se a coluna 'Data Vencimento Parcela' contém datas no formato 'DD/MM/AA'
        for valor in data_vencimento_parcela_column:
            try:
            # Tenta primeiro o formato 'DD/MM/AAAA'
                datetime.strptime(str(valor), '%d/%m/%Y')
            except ValueError:
                try:
                # Se falhar, tenta o formato padrão do Pandas 'YYYY-MM-DD HH:MM:SS'
                    datetime.strptime(str(valor), '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                    # Se falhar novamente, tenta o formato com ano de dois dígitos 'DD/MM/YY'
                        datetime.strptime(str(valor), '%d/%m/%y')
                    except ValueError:
                    # Se ainda falhar, levanta a exceção
                        raise ValidationError('A coluna Data Vencimento deve conter datas no formato DD/MM/AAAA ou YYYY-MM-DD HH:MM:SS ou DD/MM/YY.')
            
    def validate_num_parcela(self, num_parcela_column):
        # Validar se a coluna 'Num da Parcela' contém apenas números
        for valor in num_parcela_column:
            if not pd.isna(valor) and not pd.to_numeric(valor, errors='coerce'):
                raise ValidationError('A coluna Num da Parcela deve conter apenas números.')
            
    def validate_valor_parcela_format(self, valor_parcela_column):
        # Validar se a coluna 'Valor da Parcela' está no formato de valor com duas casas decimais
        for valor in valor_parcela_column:
            try:
                valor_float = float(valor)
                if not 0 <= valor_float <= 99999999.99:
                    raise ValidationError('A coluna Valor da Parcela deve estar no formato de valor com duas casas decimais.')
            except ValueError:
                raise ValidationError('A coluna Valor da Parcela deve estar no formato de valor com duas casas decimais.')
     
            
    def validate_data_pagamento_parcela_column(self, data_pagamento_parcela_column):
        for valor in data_pagamento_parcela_column:
            try:
                # Tenta primeiro o formato 'DD/MM/AAAA'
                datetime.strptime(str(valor), '%d/%m/%Y')
            except ValueError:
                try:
                    # Se falhar, tenta o formato padrão do Pandas 'YYYY-MM-DD HH:MM:SS'
                    datetime.strptime(str(valor), '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        # Se falhar novamente, tenta o formato com ano de dois dígitos 'DD/MM/YY'
                        datetime.strptime(str(valor), '%d/%m/%y')
                    except ValueError as e:
                        # Adicione uma mensagem de impressão para depurar
                        print(f'Erro ao validar formato de data para valor {valor}: {e}')
                        raise ValidationError('A coluna Data Pagamento Parcela deve conter datas no formato DD/MM/AAAA ou YYYY-MM-DD HH:MM:SS ou DD/MM/YY.')

            
    def validate_valor_pagamento_format(self, valor_pagamento_column):
        for valor in valor_pagamento_column:
            try:
                valor_float = float(valor)
                if not 0 <= valor_float <= 99999999.99:
                    raise ValidationError('A coluna Valor do Pagamento deve estar no formato de valor com duas casas decimais.')
            except ValueError:
                raise ValidationError('A coluna Valor do Pagamento deve estar no formato de valor com duas casas decimais.')
           
            
    def validate_taxa_mensal_efetiva(self, taxa_mensal_efetiva_column):
        for valor in taxa_mensal_efetiva_column:
            try:
                valor_float = float(valor)
                if not 0 <= valor_float <= 999999.99:
                    raise ValidationError('A coluna Taxa Mensal Efetiva deve ser um numeral com duas casas decimais.')
            except ValueError:
                raise ValidationError('A coluna Taxa Mensal Efetiva deve ser um numeral com duas casas decimais.')

    def validate_taxa_mensal_nominal(self, taxa_mensal_nominal_column):
        for valor in taxa_mensal_nominal_column:
            try:
                valor_float = float(valor)
                if not 0 <= valor_float <= 999999.99:
                    raise ValidationError('A coluna Taxa Mensal Nominal deve ser um numeral com duas casas decimais.')
            except ValueError:
                raise ValidationError('A coluna Taxa Mensal Nominal deve ser um numeral com duas casas decimais.')
            
    def validate_valor_mora(self, valor_mora_column):
        for valor in valor_mora_column:
            try:
                valor_float = float(valor)
                if not 0 <= valor_float <= 99999999.99:
                    raise ValidationError('A coluna Valor Mora deve ser um numeral com duas casas decimais.')
            except ValueError:
                raise ValidationError('A coluna Valor Mora deve ser um numeral com duas casas decimais.')
            
    def validate_valor_multa(self, valor_multa_column):
        for valor in valor_multa_column:
            try:
                valor_float = float(valor)
                if not 0 <= valor_float <= 99999999.99:
                    raise ValidationError('A coluna Valor Multa deve ser um numeral com duas casas decimais.')
            except ValueError:
                raise ValidationError('A coluna Valor Multa deve ser um numeral com duas casas decimais.')
            
    def validate_valor_principal(self, valor_principal_column):
        for valor in valor_principal_column:
            try:
                valor_float = float(valor)
                if not 0 <= valor_float <= 99999999.99:
                    raise ValidationError('A coluna Valor Principal deve ser um numeral com duas casas decimais.')
            except ValueError:
                raise ValidationError('A coluna Valor Principal deve ser um numeral com duas casas decimais.')
            
    def validate_valor_juros(self, valor_juros_column):
        for valor in valor_juros_column:
            try:
                valor_float = float(valor)
                if not 0 <= valor_float <= 99999999.99:
                    raise ValidationError('A coluna Valor de Juros deve ser um numeral com duas casas decimais.')
            except ValueError:
                raise ValidationError('A coluna Valor de Juros deve ser um numeral com duas casas decimais.')
            
    def validate_ultima_data_restricao_column(self, ultima_data_restricao_column):
        for valor in ultima_data_restricao_column:
            try:
                # Tenta primeiro o formato 'DD/MM/AAAA'
                datetime.strptime(str(valor), '%d/%m/%Y')
            except ValueError:
                try:
                    # Se falhar, tenta o formato padrão do Pandas 'YYYY-MM-DD HH:MM:SS'
                    datetime.strptime(str(valor), '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                        # Se falhar novamente, tenta o formato com ano de dois dígitos 'DD/MM/YY'
                        datetime.strptime(str(valor), '%d/%m/%y')
                    except ValueError as e:
                        # Adicione uma mensagem de impressão para depurar
                        print(f'Erro ao validar formato de data para valor {valor}: {e}')
                        raise ValidationError('A coluna Ultima Data Restrição deve conter datas no formato DD/MM/AAAA ou YYYY-MM-DD HH:MM:SS ou DD/MM/YY.')
            
    def validate_inclusao_restricao_column(self, inclusao_restricao_column):
        for valor in inclusao_restricao_column:
            if valor not in [0, 1]:
                raise ValidationError('A coluna Inclusão Restrição deve conter valores booleanos. 0 para Não, 1 para Sim')

            
    def validate_orgao_restricao_column(self, orgao_restricao_column, inclusao_restricao_column):
        for orgao, inclusao_restricao in zip(orgao_restricao_column, inclusao_restricao_column):
            if inclusao_restricao == 1:
                if not isinstance(orgao, str):
                    raise ValidationError('A coluna Órgão de Restrição deve conter apenas texto quando Inclusão Restrição for igual a 1.')

            
    def validate_protesto_column(self, protesto_column):
        for valor in protesto_column:
            if valor not in [0, 1]:
                raise ValidationError('A coluna Protesto deve conter valores booleanos. 0 para Não, 1 para Sim')
            
    def validate_data_protesto_column(self, data_protesto_column, protesto_column):
        for valor in data_protesto_column:
            try:
            # Tenta primeiro o formato 'DD/MM/AAAA'
                datetime.strptime(str(valor), '%d/%m/%Y')
            except ValueError:
                try:
                # Se falhar, tenta o formato padrão do Pandas 'YYYY-MM-DD HH:MM:SS'
                    datetime.strptime(str(valor), '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    try:
                    # Se falhar novamente, tenta o formato com ano de dois dígitos 'DD/MM/YY'
                        datetime.strptime(str(valor), '%d/%m/%y')
                    except ValueError:
                    # Se ainda falhar, levanta a exceção
                        raise ValidationError('A coluna Data Protesto deve conter datas no formato DD/MM/AAAA ou YYYY-MM-DD HH:MM:SS ou DD/MM/YY caso a coluna Protesto seja igual a 1.')
                
    def validate_credito_financiado_column(self, credito_financiado_column):
        for valor in credito_financiado_column:
            if not pd.isna(valor):
                try:
                    valor_float = float(valor)
                    if not 0 <= valor_float <= 99999999.99:
                        raise ValidationError('A coluna Credito Financiado deve estar no formato de valor com duas casas decimais.')
                except ValueError:
                    raise ValidationError('A coluna Credito Financiado deve estar no formato de valor com duas casas decimais.')
                
    def validate_credito_liberado_column(self, credito_liberado_column):
        for valor in credito_liberado_column:
            if not pd.isna(valor):
                try:
                    valor_float = float(valor)
                    if not 0 <= valor_float <= 99999999.99:
                        raise ValidationError('A coluna Credito Liberado deve estar no formato de valor com duas casas decimais.')
                except ValueError:
                    raise ValidationError('A coluna Credito Liberado deve estar no formato de valor com duas casas decimais.')

class UploadInclusaoForm(forms.Form):
    documento_inclusao = forms.FileField(
        required=False, 
        validators=[ExcelFileValidator()],
        widget=forms.ClearableFileInput(attrs={'class': 'form-control'})
    )

    def clean_documento_inclusao(self):
        documento_inclusao = self.cleaned_data.get('documento_inclusao')
        ExcelFileValidator()(documento_inclusao)
        return documento_inclusao

class UploadExclusaoForm(forms.Form):
    documento_exclusao = forms.FileField(
        required=False, 
        validators=[ExcelFileValidator()],
        widget=forms.ClearableFileInput(attrs={'class': 'form-control'})
        )

    def clean_documento_exclusao(self):
        documento_exclusao = self.cleaned_data.get('documento_exclusao')
        ExcelFileValidator()(documento_exclusao)
        return documento_exclusao
    
@deconstructible
class DevedorFileValidator:
    def __call__(self, value):
        if not value:
            # Se o arquivo não for fornecido, não há nada a validar
            return

        if not value.name.endswith('.xlsx'):
            raise ValidationError('O arquivo deve ter a extensão .xlsx.')

        try:
            wb = load_workbook(value)
        except Exception as e:
            raise ValidationError(f'Erro ao carregar o arquivo Excel: {e}')

        value.seek(0)

        try:
            df = pd.read_excel(value)

            # Verificando colunas obrigatórias
            required_columns = ['nome', 'cpf', 'endereco', 'telefone', 'email']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                raise ValidationError(f'O arquivo não contém as colunas obrigatórias: {", ".join(missing_columns)}.')

            if df[required_columns].dropna().empty:
                raise ValidationError('O arquivo enviado não segue os padrões fornecidos pelo arquivo de exemplo')

            self.validate_nome_column(df['nome'])
            self.validate_cpf_column(df['cpf'])
            self.validate_telefone_column(df['telefone'])

        except pd.errors.ParserError as e:
            raise ValidationError(f'Erro ao analisar o arquivo Excel: {e}')

        value.seek(0)  # Volta ao início do arquivo para que possa ser salvo no modelo

    def validate_nome_column(self, nome_column):
        # Expressão regular para verificar se a string contém apenas letras e espaços
        regex = re.compile('^[a-zA-Z ]+$')

        for valor in nome_column:
            if not regex.match(str(valor)):
                raise ValidationError('A coluna Nome deve conter apenas letras e espaços.')

    def validate_cpf_column(self, cpf_column):
        # Validar se a coluna 'CPF' contém CPFs no formato xxx.xxx.xxx-xx
        for cpf_str in cpf_column:
            if not self.validar_cpf(cpf_str):
                raise ValidationError('A coluna CPF deve conter um CPF válido e no formato XXX.XXX.XXX-XX.')

    def validar_cpf(self, cpf):
        # Remover caracteres não numéricos
        cpf_numerico = ''.join(char for char in str(cpf) if char.isdigit())

        # Verificar se o CPF tem 11 dígitos
        if not cpf_numerico.isdigit() or len(cpf_numerico) != 11:
            return False

        # Calcular os dígitos verificadores
        cpf_digits = [int(digit) for digit in cpf_numerico]

        soma1 = sum(digit * weight for digit, weight in zip(cpf_digits[:9], range(10, 1, -1)))
        resto1 = soma1 % 11
        digito_verificador1 = 0 if resto1 < 2 else 11 - resto1

        soma2 = sum(digit * weight for digit, weight in zip(cpf_digits[:10], range(11, 1, -1)))
        resto2 = soma2 % 11
        digito_verificador2 = 0 if resto2 < 2 else 11 - resto2

        # Verificar se os dígitos verificadores estão corretos
        if cpf_digits[9] != digito_verificador1 or cpf_digits[10] != digito_verificador2:
            return False

        return True

    def validate_telefone_column(self, telefone_column):
        # Validar se a coluna 'telefone' contém números e está nos formatos (XX)XXXX-XXXX ou (XX)XXXXX-XXXX
        for telefone_str in telefone_column:
            telefone_digits = ''.join(char for char in str(telefone_str) if char.isdigit())

            if not telefone_digits.isdigit() or (len(telefone_digits) != 10 and len(telefone_digits) != 11):
                raise ValidationError(
                    'A coluna Telefone deve conter apenas números e estar no formato (XX)XXXX-XXXX ou (XX)XXXXX-XXXX.')

            if len(telefone_digits) == 10:
                formatted_telefone = f'({telefone_digits[:2]}){telefone_digits[2:6]}-{telefone_digits[6:]}'
            else:
                formatted_telefone = f'({telefone_digits[:2]}){telefone_digits[2:7]}-{telefone_digits[7:]}'

            if formatted_telefone != str(telefone_str):
                raise ValidationError(
                    'A coluna Telefone deve estar no formato (XX)XXXX-XXXX ou (XX)XXXXX-XXXX.')

class InclusaoDevedorForm(forms.Form):
    documento_inclusao_devedor = forms.FileField(
        required=False,
        validators=[DevedorFileValidator()],
        widget=forms.ClearableFileInput(attrs={'class': 'form-control'})
    )

    def clean_documento_inclusao_devedor(self):
        documento_inclusao_devedor = self.cleaned_data.get('documento_inclusao_devedor')
        DevedorFileValidator()(documento_inclusao_devedor)
        return documento_inclusao_devedor
    
class ExclusaoDevedorForm(forms.Form):
    documento_exclusao_devedor = forms.FileField(
        required=False,
        validators=[DevedorFileValidator()],
        widget=forms.ClearableFileInput(attrs={'class': 'form-control'})
    )

    def clean_documento_exclusao_devedor(self):
        documento_exclusao_devedor = self.cleaned_data.get('documento_exclusao_devedor')
        DevedorFileValidator()(documento_exclusao_devedor)
        return documento_exclusao_devedor  
      
class CarteiraForm(forms.ModelForm):
    class Meta:
        model = Carteira
        fields = ['nome']
        
class CreateCarteiraForm(forms.Form):
    nome_carteira = forms.CharField(max_length=255, label='Nome da Carteira')
     
class AcordoForm(forms.ModelForm):
    class Meta:
        model = Acordo
        fields = ['cpf', 'contratos', 'email', 'telefone', 'num_parcelas_acordo', 'valor_parcela', 'dia_vencimento_primeira_parcela', 'dia_vencimento_proxima_parcela', 'valor_do_acordo']
